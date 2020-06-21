import openpyxl as xl  # giving shortforms for cleaner code
from openpyxl.chart import BarChart, Reference  # used for importing charts in sheet


def fun_transaction(filename):
    # loading workbook and sheets
    wb = xl.load_workbook(filename)
    sheet = wb["Sheet1"]
    print(".........................................................")

    # print value from sheet of specific cell
    # cell = sheet["a1"]  # both statements mean same value
    cell = sheet.cell(1, 1)  # 1st row, 1st column
    print(cell.value)
    print("...........................................................")

    # to print number of rows
    for i in range(1, sheet.max_row + 1):
        print(i)
    print("...........................................................")

    # to print values of 3rd row
    for i in range(2, sheet.max_row + 1):
        cell = sheet.cell(i, 3)  # i = row(iterates from 2 3 4 and column is 3rd)
        print(cell.value)

        # to add a column with some defines operations
        correct_price = cell.value * 0.9
        correct_cell = sheet.cell(i, 4)
        correct_cell.value = correct_price

    # adding charts in sheet using 4th column only
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)   # assigning values for chart
    chart = BarChart()
    chart.add_data(values)  # giving chart with values to be displayed
    sheet.add_chart(chart, "a6")    # commanding to add chart to the sheet with coordinates

    wb.save(filename)  # correct prices and chart updated
